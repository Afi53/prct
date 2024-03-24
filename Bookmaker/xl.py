import openpyxl
from datetime import datetime
import win32com.client as win32
import os

def chek_open(my_wb):
    try:
        xl = win32.gencache.EnsureDispatch('Excel.Application')
        if xl.Workbooks.Count > 0:
#            my_wb = 'Фонбет.xlsx'
            if any(i.Name == my_wb for i in xl.Workbooks):
                dir_path = 'C:\\Users\Professional\prct\Bookmaker'
                patch = os.path.join(dir_path, my_wb)
                wb = xl.Workbooks.Open(patch)
                wb.Visible = False
                wb.Save()
                wb.Close()
    except Exception as ex:
        print(ex)


def turn():
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx')
    sh_2 = wa['Словари']
    l_turn = []
    for col in sh_2.iter_cols(min_row=1, max_col=1, values_only=True):
        for cell in col:
            if not cell is None:
                l_turn.append(cell)  # список турниров
    return l_turn

def load_matches(match,fp,sh):
    wb = openpyxl.load_workbook(fp)
    sh_2=wb[sh]
    for m in match:
        sh_2.append(list(vars(m).values()))
    wb.save(fp)

def get_url():
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx')
    sh_2 = wa['Словари']
    d_url={}  # url Flashscore
    for row in sh_2.iter_rows(min_row=1,max_row=115,max_col=3,values_only=True):
        if not row[1] is None:
            d_url[row[0]]=row[1]
    return d_url

def get_url_ls():
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx')
    sh_1 = wa["Ставки"]
    m_row_h = len([cell for cell in sh_1['K'] if not cell.value is None]) # матчи без рез
    s_lig = sorted(set([row[0].value for row in sh_1.iter_rows(min_row=m_row_h + 1, max_col=1)])) # множество лиг
    sh_2 = wa['Словари']
    l_url=[]
    for row in sh_2.iter_rows(min_row=1,max_row=95, max_col=9,values_only=True):
        if not row[4] is None:  # отбор по стратегии
            if row[0] in s_lig:
                l_url.append(row)
    return l_url

def d_comand():
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx')
    sh_2 = wa['Словари']
    d_komand= {}    # словарь команд для Фонбет
    for row in sh_2.iter_rows(min_row=1, min_col=11, max_col=15):
        if row[2].value is None:
            d_komand[row[0].value]=row[1].value
        else:
            d_komand[row[0].value] = [row[1].value,row[2].value]
    return d_komand

def d_liges_file_name():
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx')
    sh_2 = wa['Словари']
    d_liges= {}    # словарь команд для Фонбет
    for row in sh_2.iter_rows(min_row=1, min_col=1, max_col=3):
        if not row[2].value is None:
            d_liges[row[2].value]=row[0].value
    return d_liges

def load_l_turn(turn):
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\WR.xlsx')
    sh_1=wa['Лист1']
    try:
       for row, t in enumerate(turn,start=1):
            sh_1.cell(row,column=1).value=t
    except Exception as ex:
        print(ex)
    wa.save('C:\\Users\Professional\prct\Bookmaker\WR.xlsx')
    k=1
    for i in turn:
        sh_1.cell(k,1).value=i
        k+=1
    wa.save('C:\\Users\Professional\prct\Bookmaker\WR.xlsx')

def set_liges(sh_name):
    # выгрузка списка матчей для поиска результатов
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx')
    sh_1=wa[sh_name]
    m_row_h = len([cell for cell in sh_1['K'] if not cell.value is None])
    l=sorted(set([row[0].value for row in sh_1.iter_rows(min_row=m_row_h + 2, max_col=1)]))
    print(l)

def stat():
    ws = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx')
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx',data_only=True)
    sh_1=wa["Ставки"]
    l_matches=[]
    for row in sh_1.iter_rows(min_row=2,max_col=22,values_only=True):
        if not row[10] is None:
            dat = datetime.strptime(row[1], "%d.%m.%Y").date()
            l_matches.append([row[0],dat,row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21]]                 )
    l_m_sort=sorted(l_matches,key=lambda x: x[1] ) # сортировка по дате

    sh_3=wa['Рез']

    def l_rez_lig(a, b,k_match):
        # отбор лиг с лучшими результатами
        l_rez = []
        i=0
        for row in sh_3.iter_rows(min_row=2, max_row=92, min_col=1, values_only=True):
           if row[k_match] > 66 and row[b] > 5 and row[13]!='нм':   # игры и ср результат
               l_rez.append([row[0], row[k_match], row[2], row[a], row[b]])
               i+=1
        l_rez_sort = sorted(l_rez, key=lambda x: x[4], reverse=True)[:10]
        return l_rez_sort

    def add_l_sp(l_cp,a):
# a - столбец суммирования списка матчей
        l_match_tur=[m for m in l_cp if m[2] is None]
    # отбор лиг без количесвва матчей в туре
        if any(l_match_tur)==True:
            print('нет количества матчей в туре')
            print([m[0] for m in l_match_tur])
            exit()
        for row in l_cp:
            print(row)
            l_m_lig = [m for m in l_m_sort if m[0] == row[0]]
            r= len(l_m_lig)-row[1]
            l_m_lig_sr=l_m_lig[r:]
 # список матче с результатом
            n = len(l_m_lig_sr)
            m_tur = row[2]  # матчей в туре
            if n>30*m_tur:
                l_m_lig_sr=l_m_lig_sr[-(30*m_tur):]
                k=0
            else:
                k=n%m_tur
            for i in range(k, n + 1, m_tur):
                l_otb = l_m_lig_sr[:i]
                row.append(sum([m[a] for m in l_otb])) # добавляются результаты в список лиги
        return l_cp

    def load_sh(name_sheet,l_rez):
        sh_8s = ws[name_sheet]
        for i in range(2,sh_8s.max_row+1):
            for j in range(1,sh_8s.max_column+1):
                cell=sh_8s.cell(row=i,column=j)
                cell.value=None
        for row_num, row_data in enumerate(l_rez):
            for col_num, col_data in enumerate(row_data):
                sh_8s.cell(row_num + 2, col_num + 1, col_data)

    def strateg():
        l_rez_rmk_sort = l_rez_lig(3,4,1)    # РМК
        l_cp_m = add_l_sp(l_rez_rmk_sort,2)
        load_sh('РМК', l_cp_m)

        l_rez_rbk_sort=l_rez_lig(5,6,1)      # РБК
        l_cp_b=add_l_sp(l_rez_rbk_sort,3)
        load_sh('РБК',l_cp_b)

        l_rez_rn_sort=l_rez_lig(7,8,1)      # РН
        l_cp_b=add_l_sp(l_rez_rn_sort,6)
        load_sh('РН',l_cp_b)

        l_rez_r1_sort=l_rez_lig(9,10,1)      # Р1
        l_cp_b=add_l_sp(l_rez_r1_sort,7)
        load_sh('Р1',l_cp_b)

        l_rez_r2_sort=l_rez_lig(11,12,1)      # Р2
        l_cp_b=add_l_sp(l_rez_r2_sort,8)
        load_sh('Р2',l_cp_b)

        l_rez_rmkn_sort=l_rez_lig(16,17,15)      # РМКН
        l_cp_b=add_l_sp(l_rez_rmkn_sort,4)
        load_sh('РМКН',l_cp_b)

        l_rez_rbkn_sort=l_rez_lig(18,19,15)      # РБКН
        l_cp_b=add_l_sp(l_rez_rbkn_sort,5)
        load_sh('РБКН',l_cp_b)

        l_rez_rb12_sort = l_rez_lig(20, 21, 15)  # Р12
        l_cp_b = add_l_sp(l_rez_rb12_sort, 9)
        load_sh('Р12', l_cp_b)

        l_rez_rn1_sort = l_rez_lig(22, 23, 15)  # РН1
        l_cp_b = add_l_sp(l_rez_rn1_sort, 10)
        load_sh('РН1', l_cp_b)

        l_rez_rn2_sort = l_rez_lig(24, 25, 15)  # РН2
        l_cp_b = add_l_sp(l_rez_rn2_sort, 11)
        load_sh('РН2', l_cp_b)

    def lig_itog(l_names_lig):
        l_var = [['РМК', 2], ['РБК', 3], ['РН', 6], ['Р1', 7], ['Р2', 8], ['РМКН', 4], ['РБКН', 5], ['Р12', 9],
                 ['РН1', 10], ['РН2', 11]]
        l_rez = []
        for name in l_names_lig:
            for row in sh_3.iter_rows(min_row=2, max_row=92, min_col=1, values_only=True):
               if row[0] ==name:
                   l_rez.append(row)
            l_m_lig=[m for m in l_m_sort if m[0]==name]
#            r1= len(l_m_lig) - l_rez[-1][1]
            r= l_rez[-1][1] - l_rez[-1][15]
            l_m_lig_sr = l_m_lig[r:]
            n = len(l_m_lig_sr)
            m_tur = l_rez[-1][2]
            for i in range(10):
                l_r=[]
                l_r.append(l_var[i][0]) # стратегия
                a=l_var[i][1] # номер столбца с результатами
                for k in range(n % m_tur, n + 1, m_tur):
                    l_otb = l_m_lig_sr[:k]
                    l_r.append(sum([m[a] for m in l_otb]))
                l_rez.append(l_r)
        load_sh('РЛ', l_rez)

    l_names_lig=['Англия. Премьер-Лига','Германия. Бундеслига','Италия. Серия А','Испания. Примера дивизион','Португалия. Премьер-Лига']
#    lig_itog(l_names_lig)
    strateg()

    ws.save('Фонбет.xlsx')
    ws.save('D:\\Мега\Documents and Settings\Букм\Фонбет.xlsx')

#stat()