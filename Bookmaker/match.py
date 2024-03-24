from bs4 import BeautifulSoup
import datetime
import re
import openpyxl

class Match:
    def __init__(self,liga,data,k_1,k_2,st_1,st_n,st_2,st_dv1,st_dv12,st_dv2): #,g1,g2,\
#                 rez_rmk,rez_rbk,rez_rmkn, rez_rbkn,rez_n,rez_r1,rez_r2, rez_r12,rez_rn1,rez_rn2):  #
        self.liga = liga
        self.data=data
        self.k_1=k_1
        self.k_2=k_2
        self.st_1=st_1
        self.st_n=st_n
        self.st_2=st_2
        self.st_dv1=st_dv1
        self.st_dv12=st_dv12
        self.st_dv2=st_dv2

def transform_date(date):
    d_months = {'января': '01','янв':'01', 'февраля': "02", 'фев': "02",'марта':'03', 'мар':'03',
        'апр':'04','апреля':'04', 'мая':'05', 'июня':'06',
        'июля':'07', 'августа':'08', 'сентября':'09', 'октября':'10', 'ноября':'11', 'декабря':'12'}
    def nul_md(a):   #дата и месяц <10
        a=int(a)
        if a<10:
            rez=f'0{str(a)}'
        else:
            rez=str(a)
        return rez

    if date.find('-е')!=-1:
        l_d = date.replace('-е', '').split()
        try:
            mont = nul_md(d_months[l_d[1]])
        except Exception as ex:
            print(l_d)
        day = nul_md(l_d[0])
        year = datetime.date.today().year  # год
        data_match = f'{(day)}.{(mont)}.{year}'
    else:
        d = date.split()
        if d[0].strip()=="Сегодня":
            year = datetime.date.today().year # год
            mont = datetime.date.today().month
            day = datetime.date.today().day
            data_match = f'{nul_md(day)}.{nul_md(mont)}.{year}'
        elif d[0].strip()=="Завтра":
            year=str((datetime.date.today() + datetime.timedelta(days=1)).year)
            mont = (datetime.date.today()+ datetime.timedelta(days=1)).month
            day = (datetime.date.today()+ datetime.timedelta(days=1)).day
        else:
            year = str((datetime.date.today() + datetime.timedelta(days=2)).year)
            try:
                mont= int(d_months[d[1]])
            except Exception:
                mont=3
                print('ошибка в месяяце')
            day=int(d[0])
        data_match = f'{nul_md(day)}.{nul_md(mont)}.{year}'
    return data_match

def tomorrow_data():
    datetime.date.today().day
    d=datetime.date.today().day
    m=datetime.date.today().month
    l_month_1=[1,3,5,7,8,10,12]
    l_month_2=[4.6,9,11]
    if d<29 or (d==29 and m!=2) or(d==30 and (m in l_month_1)):
        return d+1
    else:
        return 1

def d_liges():
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx')
    sh_2 = wa['Словари']
    d_liges= {}    # словарь лиг для Фонбет
    for row in sh_2.iter_rows(min_row=120, min_col=1, max_col=2,values_only=True):
        if not row[0] is None:
            d_liges[row[0].strip()]=row[1].strip()
    return d_liges

def get_data_LS():
    d_ligs=d_liges()
    with open("index_LS.html","r",encoding='utf-8') as file:
        data=file.read()
    soup=BeautifulSoup(data,'lxml')
    now = datetime.date.today().strftime('%d''.''%m''.''%Y')
#    print(now)
    l_matches = []
    liges=soup.find_all('div',class_='content__inner-wrapper-c1fc01')
    n=len(liges)
    for lig in liges:
        tur=lig.find('h1').text.replace('Ставки на ','').replace('. ',',').strip()
        print(tur)
        if tur in d_ligs.values():
            liga=tur
        else:
            try:
                liga = d_ligs[tur]
            except KeyError:
                print(f'{tur} ошибка словаря ЛС')
                continue
        section=lig.find('div', class_='events-proposed__wrapper-events-f8fbd6')
        matches=section.find_all('div',class_='bui-event-row-dfbc70')
        n=len(matches)
        for m in matches:
            d=m.find('div',class_='bui-event-row__date-c9e228').text  # дата
            if d=='\n':
                continue
            date=transform_date(d)
            if now!=date:
                if tomorrow_data()==int(date.split(".")[0]):
                    h_r=m.find('div',class_='bui-event-row__column-be08c1 bui-event-row__column_meta-d53051').text
                    hour=int(h_r.replace('\n','').strip().split(':')[0])
                    if hour>12:
                        continue
                else:
                    continue
            kom=m.find('div',class_='bui-event-row__info-c6ed29').find_all('span',class_='bui-commands__command-d517c1')
            k_1=kom[0].text.strip()
            if k_1.split()[-1]=='Итоги':
                break
            k_2=kom[1].text.strip()
            st=m.find_all('span',class_='bui-outcome__title-1ebb32')
            st_1=float(st[0].find_next_sibling('span').text.strip().replace(',','.'))
            st_n=float(st[1].find_next_sibling('span').text.strip().replace(',','.'))
            st_2=float(st[2].find_next_sibling('span').text.strip().replace(',','.'))
#            st_dv1,st_dv2,st_dv12=None,None,None
#            m = Match(liga, date, k_1, k_2, st_1, st_n, st_2,st_dv1,st_dv2,st_dv12)
            m = Match(liga, date, k_1, k_2, st_1, st_n, st_2,None,None,None)
            l_matches.append(m)
    l_matches_BS= get_data_BS(d_ligs)
    l_matches_M=get_data_Mar(d_ligs)
    join_Lig(l_matches, l_matches_BS,l_matches_M)
    fp='C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx'
    sh='ЛС_рез'
    return l_matches,fp,sh

def conv_st(stavka):
    a=stavka.replace(',','.')
    b=float(a)
    return b

def get_data_BS(d_ligs):
    with open("index_BS.html","r",encoding='utf-8') as file:
        data=file.read()
    soup=BeautifulSoup(data,'lxml')
    l_matches_BS = []
    liges=soup.find_all('div',class_='line-champ__header-name')
    n=len(liges)
    print('BS')
#    d_ligs['Чили. Примера Дивизион']=['Чили.Серия а']
    for lig in liges:
        tur=lig.find('h2').text.replace('Футбол. ','').strip()
        if tur.find('Лига Европы УЕФА')!=-1:
            tur='Лига Европы УЕФА'
        elif tur.find('Лига Конференций УЕФА')!=-1:
            tur='Лига Конференций УЕФА'
        elif tur[-1]=='.':
            tur=(tur[:(len(tur)-1)]).strip()
        print(tur,type(tur))
        if tur in d_ligs.values():
            liga=tur
        else:
            try:
                liga = d_ligs[tur]
            except KeyError:
                print(f'{tur} ошибка словаря')
                continue
        dt = lig.find_next('div', class_='line-champ__date').text.strip()
#        print(dt)
        date = transform_date(dt)
        matches = lig.find_all_next('app-line-event-unit',{'_nghost-desktop-ng-cli-c173': ""})  # ('app-line-event-unit')  #
        т = len(matches)
        for match in matches:
            komands = match.find('span', class_=re.compile('line-event__name-teams line-event__name-teams'))
#            for name in komands:
            k_1=(komands.find_all('b')[0].text).strip()
            k_2=(komands.find_all('b')[1].text).strip()
 #           print(k_1,k_2)
            st = match.find_all('button', {'class': 'line-event__main-bets-button'})
            st_1=float(st[0].text.strip())
            st_n=float(st[1].text.strip())
            st_2=float(st[2].text.strip())
            st_dv = match.find_all('div', {'data-first-index': "WXm"})  # {'data-section-id': re.compile(r"137")})
            d_st_dv={'1X':0,'12':0,'X2':0}
            for s in st_dv:
                st_name = s.text.strip()[:2]
                for key,value in d_st_dv.items():
                    if st_name==key:
                        try:
                            d_st_dv[key]= float(s.find_all('button')[0].text.strip())
                        except IndexError:
                            pass
                        break
            m = Match(liga, date, k_1, k_2, st_1, st_n, st_2,d_st_dv['1X'],d_st_dv['12'],d_st_dv['X2'])
            l_matches_BS.append(m)
            tag = match.find_next_sibling()
            if tag is None:
                break
            tag_str = str(lig)[12:27]
    return l_matches_BS           #,fp,sh

def find_kom(value,l_kom):
    # поиск по индексу в списке списков
    for list in l_kom:
        if value in list:
            return l_kom.index(list)
    return -1

def l_komand():
    # cписок списков названий команд
    wa = openpyxl.load_workbook('C:\\Users\Professional\prct\Bookmaker\Фонбет.xlsx')
    sh_2 = wa['Словари']
    l_kom=[]
    for row in sh_2.iter_rows(min_row=1, min_col=11, max_col=15):
        l_list=[]
        for i in range(4):
            if not row[i].value is None:
                l_list.append(row[i].value)
        l_kom.append(l_list)
    return l_kom

def join_Lig(l_matches_LS,l_matches_BS,l_matches_M):
    l_kom=l_komand()
    l_join = []
    def join_app(a:Match,l_sp,bukm):
        l_join.append([None, bukm, None, None, a.st_1, a.st_n, a.st_2, a.st_dv1, a.st_dv12, a.st_dv2])
        l_sp.remove(a)

    for ls in l_matches_LS:
        l_join.append(list(vars(ls).values())[:7])

        def app_list(a: Match, l_matches_lig,bukm):
            def find_match(a_1,a_2):
                for t in [a_1,a_2]:
                    ind = find_kom(t, l_kom)  # индекс строки c назв команды в списке  названий
                    if ind!=-1:
                        for b in lm_lig:
                            if b.k_1 in l_kom[ind]:
                                join_app(b, l_matches_lig, bukm)
                                return True
                            else:
                                print(f'{bukm} {b.k_1} команды нет в строке {a.k_1}')
                                return True
                    else:
                        print(f'{bukm} {a.k_1} команды нет в списке')
            m=next((k for k in l_matches_lig if (k.k_1==a.k_1 or k.k_2==a.k_2)),None)
            if not m is None:
                join_app(m, l_matches_lig, bukm)
            else:
                lm_lig=[k for k in l_matches_lig if k.liga==a.liga]
                if any(lm_lig)==False:
                    print(f'в {bukm} нет лиги {a.liga}')
                    return
                if find_match(a.k_1,a.k_2)==True:
                    return
                for t in [a.k_1,a.k_2]:
                    ind = find_kom(t, l_kom)  # индекс строки c назв команды в списке  названий
                    if ind!=-1:
                        for b in lm_lig:
                            if b.k_1 in l_kom[ind]:
                                join_app(b, l_matches_lig, bukm)
                                break
                            else:
                                print(f'{bukm} {b.k_1} команды нет в строке {a.k_1}')
                    else:
                        print(f'{bukm} {a.k_1} команды нет в списке {a.liga}')

        app_list(ls,l_matches_BS,'BS')
        app_list(ls,l_matches_M, 'Mar')
    wa = openpyxl.load_workbook('D:\\Мега\Documents and Settings\Букм\CC.xlsx')
    sh_1 = wa['Лист1']
    sh_1.delete_rows
    for row in l_join:
        sh_1.append(row)
    wa.save('D:\\Мега\Documents and Settings\Букм\CC.xlsx')

def get_data_Mar(d_ligs):
    with open("index_M.html","r",encoding='utf-8') as file:
        data=file.read()
    soup=BeautifulSoup(data,'lxml')
    l_matches_M = []
    liges=soup.find_all('div','category-container')
    n=len(liges)
    print('Марафон')
    for lig in liges:
        t=lig.find('h2').text.replace('\n','').strip()
        tur=" ".join(t.split()).strip()
        print(tur)
        if tur.find('Итоги')!=-1:
            continue
        elif tur.find('Лига Европы UEFA')!=-1:
            tur='Лига Европы UEFA'
        elif tur.find('Лига конференций UEFA')!=-1:
            tur='Лига конференций UEFA'
        if tur in d_ligs.values():
            liga=tur
        else:
            try:
                liga = d_ligs[tur]
            except KeyError:
                print(f'{tur} ошибка словаря')
                continue
        matches=lig.find_all('div',class_='bg coupon-row')
        n=len(matches)
        for match in matches:
            kom=match["data-event-name"]
            k_1=(kom.split(' - ')[0]).strip()
            k_2=(kom.split(' - ')[1]).strip()
            d=match.find('td',class_=re.compile('date date-')).text.replace('\n','').strip()
#            print(d)
            if len(d)<6:
                date= datetime.date.today().strftime('%d''.''%m''.''%Y')
            else:
                date=transform_date(d)
            st=match.find_all('td',class_=re.compile('height-column-with'))  #'selection-link active-selection')
            st_1 = float(st[0].text.strip())
            st_n = float(st[1].text.strip())
            st_2 = float(st[2].text.strip())
            try:
                st_dv1 = float(st[3].text.strip())
            except ValueError:
                st_dv1=0
            try:
                st_dv12 = float(st[4].text.strip())
            except ValueError:
                st_dv12=0
            try:
                st_dv2 = float(st[5].text.strip())
            except ValueError:
                st_dv2=0
            m = Match(liga, date, k_1, k_2, st_1, st_n, st_2,st_dv1,st_dv12,st_dv2)
            l_matches_M.append(m)
    return l_matches_M

